
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl.styles import (
    Alignment, Font
)
import os
import datetime


class CExcelLog:
    folder_name = "scanned"

    # СТИЛЬ ШРИФТА
    FONT_HEADER = Font(
        name='Calibri',
        size=21,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    FONT_DATA = Font(
        name='Arial',
        size=18,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0
    )
    def __init__(self):
        pass

    @classmethod
    def print_log(cls, pallett_code: str, tv_sn: str, tv_fk: int) -> bool:

        folder_its_ok = False
        if os.path.isdir(f"{cls.folder_name}"):
            folder_its_ok = True
        else:
            os.mkdir(cls.folder_name + "/")

            if os.path.isdir(f"{cls.folder_name}") is True:
                folder_its_ok = True
        ###
        if folder_its_ok is True:

            cdate = datetime.datetime.now()
            #
            day = cdate.day
            month = cdate.month
            year = cdate.year

            secs = cdate.second
            hours = cdate.hour
            mins = cdate.minute

            file_name = f"pscanned_log_{day}.{month}.{year}.xlsx"

            is_file_exists = False
            file_name_with_patch = f"{cls.folder_name}/{file_name}"
            if os.path.exists(file_name_with_patch) is True:
                is_file_exists = True

            # try:
            if not is_file_exists:
                wb = Workbook()
                ws = wb.active
                ws.title = "Сканировка паллетов"
                ws.append(("Паллет №:", "SN телевизора:", "ID модели:", "Дата сканировки:"))

                # задаём ширину и фонт для шапки
                cell_range = ws['A1':'D1']
                for i in cell_range:
                    for i2 in i:
                        letter_adress = i2.coordinate

                        ws[letter_adress].font = cls.FONT_HEADER
                        ws[letter_adress].alignment = cls.alignment

                interval = get_column_interval("A", "D")
                for item in interval:
                    ws.column_dimensions[item].width = 60
            else:
                wb = load_workbook(file_name_with_patch)
                ws = wb.active

            ws.append((pallett_code, tv_sn, tv_fk, f"{hours:02}:{mins:02}:{secs:02} {day:02}:{month:02}:{year}"))

            # # Задаём фонт для столбцов с данными

            #TODO Понять как весь лист ебануть под один шрифт
            # Не получается ебануть строку, всё время она смещается вниз после стиля

            # cell_range = ws['A2':'D2500']
            # for i in cell_range:  # Можешь изменить min_row, чтобы начать с другой строки
            #     # Проходимся по каждой ячейке в строке
            #     for cell in i:
            #         # Применяем шрифт к ячейке
            #         cell.font = cls.FONT_HEADER
            #         cell.alignment = cls.alignment


            wb.save(f"{cls.folder_name}/{file_name}")
            wb.close()

            #except Exception as err:
                #print(f"Внимание! Ошибка лога: '{err}'.")
                #return False

            # if os.path.exists(f"{cls.folder_name}/{file_name}") is True:
            #     pass
            # else:
            return True




